"""Text and Excel reporting for monthly airline route exits."""

from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
import pandas as pd


def format_number(value) -> str:
    if pd.isna(value):
        return ""
    return f"{float(value):,.0f}"


def write_text_report(
    events: pd.DataFrame,
    output: Path,
    *,
    args: argparse.Namespace,
    rows_read: int,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    route_definition = (
        "directional airport pair" if args.directional else "undirected airport pair"
    )
    if args.carrier_level == "marketing":
        service_definition = (
            "BTS Marketing Carrier On-Time scheduled domestic flights; "
            "cancelled flights do not establish active service"
        )
    else:
        service_definition = (
            "all T-100 service classes"
            if args.all_service_classes
            else "T-100 service class(es) " + ", ".join(args.service_classes)
        )
    lines = [
        "AIRLINE NONSTOP ROUTE EXIT EVENTS",
        "",
        f"Carrier level: {args.carrier_level}",
        f"Route definition: {route_definition}",
        "Connecting options: ignored; only nonstop flight records are examined",
        f"Service included: {service_definition}",
        (
            "Active route-month: at least "
            f"{args.min_active_departures:g} performed departure(s)"
            + ("" if args.carrier_level == "marketing" else " and positive seats")
        ),
        f"Exit requirement: at least {args.min_absence_months} consecutive inactive months",
        f"Minimum active months in prior 12: {args.min_active_months_before_exit}",
        (
            "Minimum performed flights in prior 12: "
            f"{args.min_performed_flights_before_exit:g}"
        ),
        (
            "Competition at exit: "
            + (
                "required; at least one other airline must have a performed "
                "nonstop flight on the exact airport pair in the exit-start month"
                if args.require_competing_airline_at_exit
                else (
                    "recorded as a statistic only and not required for an event "
                    "to qualify"
                )
            )
        ),
        (
            "Route-continuation measure: consecutive months beginning in the "
            "exit-start month with at least one different airline performing a "
            "nonstop flight on the exact airport pair; ends at the first month "
            "with no other airline"
        ),
        f"Raw source rows read: {rows_read:,}",
        f"Qualifying exit episodes: {len(events):,}",
        "",
    ]
    for index, event in enumerate(events.itertuples(index=False), start=1):
        location = f"{event.airport_1}-{event.airport_2}"
        if pd.notna(event.city_1) or pd.notna(event.city_2):
            location += f" ({event.city_1} / {event.city_2})"
        airline = event.airline if pd.notna(event.airline) else "Name unavailable"
        carrier_label = (
            "Marketing carrier" if args.carrier_level == "marketing" else "Carrier"
        )
        prior_service_line = (
            (
                "   Prior 12 months: "
                f"{event.active_months_in_prior_12} active month(s), "
                f"{format_number(event.departures_in_prior_12)} performed flights"
            )
            if args.carrier_level == "marketing"
            else (
                "   Prior 12 months: "
                f"{event.active_months_in_prior_12} active month(s), "
                f"{format_number(event.departures_in_prior_12)} departures, "
                f"{format_number(event.seats_in_prior_12)} seats, "
                f"{format_number(event.passengers_in_prior_12)} passengers"
            )
        )
        lines.extend(
            [
                f"{index}. {carrier_label}: {event.carrier} - {airline}",
                (
                    "   Operating carrier(s) in prior 12 months: "
                    f"{event.operating_carriers_prior_12}"
                ),
                f"   Nonstop route: {location}",
                f"   First observed service: {event.first_observed_service_month}",
                f"   Last service before exit: {event.last_service_month}",
                f"   Exit began: {event.exit_start_month}",
                (
                    "   Other airline(s) flying the same nonstop route in the "
                    f"exit month: {event.competing_airlines_at_exit}"
                ),
                (
                    "   Continuous service by other airlines after exit: "
                    f"{event.continuous_other_airline_service_months} month(s)"
                    + (
                        " through " + event.continuous_other_airline_service_through
                        if event.continuous_other_airline_service_through
                        else ""
                    )
                    + f"; {event.other_airline_continuation_status}"
                ),
                (
                    "   Airline(s) during that continuous service: "
                    f"{event.airlines_during_continuous_service}"
                ),
                (
                    "   One-year absence verified through: "
                    f"{event.one_year_absence_verified_through}"
                ),
                (
                    f"   Return: {event.return_month}"
                    if event.return_month
                    else (
                        "   Return: none observed through "
                        f"{event.absence_observed_through}"
                    )
                ),
                (
                    f"   Observed absence: {event.observed_absence_months} month(s); "
                    f"{event.exit_status}"
                ),
                prior_service_line,
                "",
            ]
        )
    output.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def build_summary_tables(
    events: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Aggregate qualifying exit events into airline and airport summaries."""
    airline_columns = [
        "Carrier Code",
        "Airline",
        "Affected Routes",
        "Exit Events",
        "Performed Flights",
        "Routes",
    ]
    airport_columns = [
        "Airport",
        "Affected Routes",
        "Affected Airlines",
        "Exit Events",
        "Performed Flights",
        "Airlines",
        "Routes",
    ]
    if events.empty:
        return pd.DataFrame(columns=airline_columns), pd.DataFrame(
            columns=airport_columns
        )

    airline_rows = [
        {
            "Carrier Code": carrier,
            "Airline": airline if pd.notna(airline) else "Name unavailable",
            "Affected Routes": int(group["route"].nunique()),
            "Exit Events": int(len(group)),
            "Performed Flights": float(group["departures_in_prior_12"].sum()),
            "Routes": ", ".join(sorted(group["route"].unique())),
        }
        for (carrier, airline), group in events.groupby(
            ["carrier", "airline"], dropna=False, observed=True
        )
    ]
    airline_table = pd.DataFrame(airline_rows, columns=airline_columns).sort_values(
        ["Performed Flights", "Carrier Code"], ascending=[False, True]
    )

    airport_records = [
        {
            "Airport": airport,
            "Route": event.route,
            "Carrier Code": event.carrier,
            "Airline": (
                event.airline if pd.notna(event.airline) else "Name unavailable"
            ),
            "Performed Flights": float(event.departures_in_prior_12),
        }
        for event in events.itertuples(index=False)
        for airport in (event.airport_1, event.airport_2)
    ]
    airport_events = pd.DataFrame(airport_records)
    airport_rows = []
    for airport, group in airport_events.groupby("Airport", observed=True):
        airline_labels = sorted(
            {
                f"{row[0]} - {row[1]}"
                for row in group[["Carrier Code", "Airline"]].itertuples(
                    index=False, name=None
                )
            }
        )
        airport_rows.append(
            {
                "Airport": airport,
                "Affected Routes": int(group["Route"].nunique()),
                "Affected Airlines": int(group["Carrier Code"].nunique()),
                "Exit Events": int(len(group)),
                "Performed Flights": float(group["Performed Flights"].sum()),
                "Airlines": ", ".join(airline_labels),
                "Routes": ", ".join(sorted(group["Route"].unique())),
            }
        )
    airport_table = pd.DataFrame(airport_rows, columns=airport_columns).sort_values(
        ["Performed Flights", "Airport"], ascending=[False, True]
    )
    return airline_table.reset_index(drop=True), airport_table.reset_index(drop=True)


def build_event_table(events: pd.DataFrame) -> pd.DataFrame:
    """Create one machine-readable Excel row per qualifying exit event."""
    columns = [
        "Nonstop Route",
        "Airport 1",
        "Airport 2",
        "Exiting Carrier",
        "Exiting Airline",
        "Last Service Month",
        "Exit Start Month",
        "Another Airline on Route at Exit",
        "Other Airlines Flying Same Route at Exit",
        "Other Airline Count",
        "Other Airline Performed Flights in Exit Month",
        "Continuous Other-Airline Service Months After Exit",
        "Continuous Other-Airline Service Through",
        "Other-Airline Continuation Status",
        "Airlines During Continuous Service",
        "Exiting Airline Performed Flights in Prior 12 Months",
        "Operating Carriers in Prior 12 Months",
        "One-Year Absence Verified Through",
        "Return Month",
        "Observed Absence Months",
        "Exit Status",
    ]
    if events.empty:
        return pd.DataFrame(columns=columns)
    table = pd.DataFrame(
        {
            "Nonstop Route": events["route"],
            "Airport 1": events["airport_1"],
            "Airport 2": events["airport_2"],
            "Exiting Carrier": events["carrier"],
            "Exiting Airline": events["airline"].fillna("Name unavailable"),
            "Last Service Month": events["last_service_month"],
            "Exit Start Month": events["exit_start_month"],
            "Another Airline on Route at Exit": events[
                "competing_airline_count_at_exit"
            ]
            .gt(0)
            .map({True: "Yes", False: "No"}),
            "Other Airlines Flying Same Route at Exit": events[
                "competing_airlines_at_exit"
            ],
            "Other Airline Count": events["competing_airline_count_at_exit"],
            "Other Airline Performed Flights in Exit Month": events[
                "competing_departures_at_exit"
            ],
            "Continuous Other-Airline Service Months After Exit": events[
                "continuous_other_airline_service_months"
            ],
            "Continuous Other-Airline Service Through": events[
                "continuous_other_airline_service_through"
            ],
            "Other-Airline Continuation Status": events[
                "other_airline_continuation_status"
            ],
            "Airlines During Continuous Service": events[
                "airlines_during_continuous_service"
            ],
            "Exiting Airline Performed Flights in Prior 12 Months": events[
                "departures_in_prior_12"
            ],
            "Operating Carriers in Prior 12 Months": events[
                "operating_carriers_prior_12"
            ],
            "One-Year Absence Verified Through": events[
                "one_year_absence_verified_through"
            ],
            "Return Month": events["return_month"],
            "Observed Absence Months": events["observed_absence_months"],
            "Exit Status": events["exit_status"],
        },
        columns=columns,
    )
    return table.sort_values(
        ["Exit Start Month", "Exiting Carrier", "Nonstop Route"]
    ).reset_index(drop=True)


def write_excel_tables(events: pd.DataFrame, output: Path) -> None:
    """Write formatted event, airline, and airport worksheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError(
            "Excel output requires openpyxl. Install the packages in requirements.txt."
        ) from exc

    event_table = build_event_table(events)
    airline_table, airport_table = build_summary_tables(events)
    workbook = Workbook()
    workbook.remove(workbook.active)
    source_url = "https://www.transtats.bts.gov/Fields.asp?gnoyr_VQ=FGK"
    navy = "17365D"
    white = "FFFFFF"
    gray = "5B6573"
    thin_border = Border(bottom=Side(style="thin", color="B4C7E7"))

    def prepare_sheet(name: str, title: str, subtitle: str):
        sheet = workbook.create_sheet(name)
        sheet.sheet_view.showGridLines = False
        sheet["A1"] = title
        sheet["A1"].font = Font(bold=True, size=18, color=navy)
        sheet["A2"] = subtitle
        sheet["A2"].font = Font(italic=True, size=10, color=gray)
        sheet["A3"] = f"Source: {source_url}"
        sheet["A3"].hyperlink = source_url
        sheet["A3"].style = "Hyperlink"
        return sheet

    def write_frame(
        sheet,
        frame: pd.DataFrame,
        *,
        table_name: str,
        widths: list[float],
        header_height: float,
        row_height: float | None,
        wrap_all: bool = False,
    ) -> tuple[int, int]:
        header_row = 5
        for column_index, header in enumerate(frame.columns, start=1):
            cell = sheet.cell(header_row, column_index, header)
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(bold=True, color=white)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[header_row].height = header_height
        for row_index, row in enumerate(
            frame.itertuples(index=False, name=None), start=header_row + 1
        ):
            for column_index, value in enumerate(row, start=1):
                if isinstance(value, np.generic):
                    value = value.item()
                cell = sheet.cell(row_index, column_index, value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=wrap_all)
            if row_height is not None:
                sheet.row_dimensions[row_index].height = row_height
        final_row = header_row + max(len(frame), 1)
        final_column = len(frame.columns)
        if len(frame):
            end_cell = sheet.cell(final_row, final_column).coordinate
            excel_table = Table(
                displayName=table_name, ref=f"A{header_row}:{end_cell}"
            )
            excel_table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(excel_table)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
        sheet.freeze_panes = "A6"
        return final_row, final_column

    event_sheet = prepare_sheet(
        "Exit Events",
        "Nonstop Route Exit Events and Continuing Service",
        (
            f"{len(events):,} qualifying events. Other-airline service on the "
            "exact nonstop airport pair is reported as a statistic, not a filter."
        ),
    )
    event_widths = [
        16,
        12,
        12,
        14,
        24,
        17,
        17,
        18,
        58,
        16,
        24,
        22,
        22,
        32,
        62,
        25,
        58,
        22,
        16,
        18,
        32,
    ]
    event_final_row, _ = write_frame(
        event_sheet,
        event_table,
        table_name="NonstopExitEvents",
        widths=event_widths,
        header_height=42,
        row_height=None,
        wrap_all=True,
    )
    for row in range(6, event_final_row + 1):
        wrapped_text_specs = ((9, 75), (15, 80), (17, 75))
        estimated_lines = max(
            max(
                1,
                (
                    len(str(event_sheet.cell(row, column).value or ""))
                    + characters_per_line
                    - 1
                )
                // characters_per_line,
            )
            for column, characters_per_line in wrapped_text_specs
        )
        event_sheet.row_dimensions[row].height = min(
            150, max(48, estimated_lines * 15)
        )
        for column in (10, 11, 12, 16, 20):
            event_sheet.cell(row, column).number_format = "#,##0"

    summary_specs = (
        (
            "By Airline",
            "Affected Routes by Marketing Airline",
            (
                f"{len(events):,} qualifying exit events; flight totals cover each "
                "event's 12-month pre-exit window."
            ),
            airline_table,
            "AirlineExitSummary",
            [14, 25, 16, 14, 20, 85],
            105,
            (6,),
        ),
        (
            "By Airport",
            "Affected Routes and Airlines by Airport",
            (
                "Each route is attributed to both endpoint airports; flight totals "
                "therefore should not be summed across airports."
            ),
            airport_table,
            "AirportExitSummary",
            [12, 16, 17, 14, 20, 48, 70],
            72,
            (6, 7),
        ),
    )
    for (
        name,
        title,
        subtitle,
        frame,
        table_name,
        widths,
        row_height,
        wrap_columns,
    ) in summary_specs:
        sheet = prepare_sheet(name, title, subtitle)
        final_row, final_column = write_frame(
            sheet,
            frame,
            table_name=table_name,
            widths=widths,
            header_height=28,
            row_height=row_height,
        )
        for row in range(6, final_row + 1):
            for column in range(3, min(5, final_column) + 1):
                sheet.cell(row, column).number_format = "#,##0"
            for column in wrap_columns:
                sheet.cell(row, column).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
